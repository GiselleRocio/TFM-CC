"""
exp07_gurobi_scalability.py — Análisis de escalabilidad de Gurobi (offshore BAP, time-indexed)

Pregunta: ¿En qué N empieza a costar? ¿Cuándo no certifica óptimo en 300 s?

Autocontenido: genera sus propias instancias, no depende de instances.xlsx.
Formulación MILP idéntica a exp01 (assignment, no big-M).

Outputs:
  results/exp07_gurobi_scalability.xlsx   (hoja: raw_runs)
  results/exp07_walltime_scaling.png
  results/exp07_mipgap_scaling.png
  results/exp07_nodes_scaling.png
  results/exp07_summary.csv

Celdas:
  1 — Imports, configuración, helpers copiados de setup.py
  2 — _solve_gurobi_scale(N, seed) -> dict
  3 — RUN paralelo (ProcessPoolExecutor), append-safe
  4 — Detección automática de límites N*, N**, N***
  5 — Plots
  6 — Resumen textual + exp07_summary.csv
"""

# CELDA 1: IMPORTS, CONFIGURACIÓN Y HELPERS (Colab)
# ---- EDITAR SI TU CARPETA TIENE OTRO NOMBRE ---
DRIVE_TESIS_PATH = "MyDrive/TESIS"
# -----------------------------------------------

import os, sys, math, time, datetime, logging, subprocess
import importlib.util as _ilu
from pathlib import Path

from google.colab import drive
drive.mount("/content/drive", force_remount=False)

DRIVE_TESIS      = f"/content/drive/{DRIVE_TESIS_PATH}"
REPO_ROOT        = Path(DRIVE_TESIS)
EXPERIMENTS2_DIR = REPO_ROOT / "experiments2"

for p in [str(REPO_ROOT / "src"), str(REPO_ROOT)]:
    if p not in sys.path:
        sys.path.insert(0, p)

# Instalar gurobipy si no esta disponible
if _ilu.find_spec("gurobipy") is None:
    print("  instalando gurobipy...", end=" ", flush=True)
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "gurobipy"],
                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    print("listo")
else:
    print("  ok  gurobipy")

# Instalar openpyxl si no esta disponible
if _ilu.find_spec("openpyxl") is None:
    print("  instalando openpyxl...", end=" ", flush=True)
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"],
                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    print("listo")
else:
    print("  ok  openpyxl")

# Cargar credenciales Gurobi WLS: primero Colab Secrets, luego TESIS/.env
for _gk in ("GRB_WLSACCESSID", "GRB_WLSSECRET", "GRB_LICENSEID"):
    if _gk not in os.environ:
        try:
            from google.colab import userdata
            _val = userdata.get(_gk)
            if _val:
                os.environ[_gk] = _val
        except Exception:
            pass

# Fallback: leer .env si las variables aun no estan en el entorno
if not all(k in os.environ for k in ("GRB_WLSACCESSID", "GRB_WLSSECRET", "GRB_LICENSEID")):
    _dotenv_path = REPO_ROOT / ".env"
    if _dotenv_path.exists():
        for _ln in _dotenv_path.read_text().splitlines():
            _ln = _ln.strip()
            if _ln and not _ln.startswith("#") and "=" in _ln:
                _k, _, _v = _ln.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip())
        print("  ok  credenciales Gurobi cargadas desde TESIS/.env")
    else:
        print("  AVISO: credenciales Gurobi WLS no configuradas (GRB_WLSACCESSID etc).")
else:
    print("  ok  credenciales Gurobi WLS (Colab Secrets)")

import numpy as np
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(name)s | %(message)s")
logger = logging.getLogger("exp07")

# ── Configuración del experimento ────────────────────────────────────────────
SCALE_GRID  = [8, 12, 16, 20, 30, 40, 60, 80, 100, 120, 150, 175, 200]
SEEDS       = [42, 43, 44, 45, 46]
TIME_LIMIT  = 300          # segundos
MIP_GAP     = 1e-4         # 0.01%
RHO_TARGET  = 0.80
MIX_VLCC    = 0.25
R_J_DIST    = "clustered"
MIN_CD      = 0.60

RESULTS_DIR = EXPERIMENTS2_DIR / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_PATH = RESULTS_DIR / "exp07_gurobi_scalability.xlsx"
SHEET       = "raw_runs"

print(f"  ok  paths configurados")
print(f"  REPO_ROOT: {REPO_ROOT}")

# ── Helpers copiados de setup.py (autocontenido) ──────────────────────────────

def _rho_effective(nominations_df: pd.DataFrame, T: int) -> float:
    return float(nominations_df["p_j"].sum()) / T


def _n_vars_approx(n: int, T: int, vlcc_pct: float = 0.25,
                   p_std: int = 4, p_vlcc: int = 8, n_machines: int = 2) -> int:
    n_vlcc = round(n * vlcc_pct)
    n_std  = n - n_vlcc
    return (n_std * max(0, T - p_std + 1) + n_vlcc * max(0, T - p_vlcc + 1)) * n_machines


def _conflict_density(nominations_df: pd.DataFrame) -> float:
    # Fraccion de pares (j,k) con ventanas solapadas.
    r = nominations_df["r_j"].values
    d = nominations_df["d_j"].values
    n = len(r)
    if n < 2:
        return 0.0
    total     = n * (n - 1) // 2
    conflicts = sum(
        1
        for i in range(n)
        for j in range(i + 1, n)
        if r[i] < d[j] and r[j] < d[i]
    )
    return round(conflicts / total, 4)


def _t_from_rho(n: int, vlcc_pct: float, rho_target: float) -> int:
    n_vlcc  = round(n * vlcc_pct)
    n_std   = n - n_vlcc
    total_p = n_vlcc * 8 + n_std * 4
    return math.ceil(total_p / rho_target)


def _make_nominations(n: int, T: int, vlcc_pct: float,
                      r_j_distribution: str, seed: int,
                      min_conflict_density: float = 0.6,
                      collision_target: float | None = None) -> pd.DataFrame:
    p_max          = max(8, round(vlcc_pct * 8 + (1 - vlcc_pct) * 4) + 2)
    max_density_seen = 0.0

    for attempt in range(20):
        rng    = np.random.default_rng(seed + attempt)
        n_vlcc = round(n * vlcc_pct)
        n_std  = n - n_vlcc
        p_j    = np.array([8] * n_vlcc + [4] * n_std, dtype=int)
        rng.shuffle(p_j)

        r_max = max(1, T - max(p_j) - 2)
        if collision_target is not None and r_j_distribution == "clustered":
            if attempt >= 15:
                r_max = max(1, int(r_max * 0.25))
            elif attempt >= 10:
                r_max = max(1, int(r_max * 0.50))
            elif attempt >= 5:
                r_max = max(1, int(r_max * 0.75))

        if r_j_distribution == "clustered":
            n_cluster    = max(1, round(n * 0.7))
            n_spread     = n - n_cluster
            cluster_time = rng.integers(0, max(1, r_max - T // 3))
            spread_times = rng.integers(0, max(1, r_max + 1), size=max(1, n_spread // 2))
            r_cluster    = np.full(n_cluster, cluster_time)
            r_spread     = rng.choice(spread_times, size=n_spread) if n_spread > 0 else np.array([])
            r_j          = np.concatenate([r_cluster, r_spread])
        elif r_j_distribution == "uniform":
            n_batches  = max(1, n // 3)
            batch_times = rng.integers(0, max(1, r_max + 1), size=n_batches)
            r_j         = rng.choice(batch_times, size=n)
        elif r_j_distribution == "random":
            n_batches  = max(1, n // 3)
            batch_times = rng.integers(0, max(1, r_max // 2), size=n_batches)
            r_j         = rng.choice(batch_times, size=n)
        elif r_j_distribution == "bimodal":
            half  = n // 2
            peak1 = max(0, T // 6)
            peak2 = max(0, T // 2)
            r_j   = np.concatenate([np.full(half, peak1), np.full(n - half, peak2)])
        else:
            raise ValueError(f"r_j_distribution desconocida: {r_j_distribution!r}")

        r_j = np.sort(r_j.astype(int))

        t_free     = 0
        fits_in_T  = True
        for arr, proc in zip(r_j, p_j):
            start = max(t_free, arr)
            if start + proc > T:
                fits_in_T = False
                break
            t_free = start + proc
        if not fits_in_T:
            continue

        slack_vals = rng.integers(1, 4, size=n)
        d_j = np.minimum(r_j + p_j + slack_vals, T)
        d_j = np.maximum(d_j, r_j + p_j)

        noms_temp       = pd.DataFrame({"r_j": r_j, "d_j": d_j, "p_j": p_j})
        current_density = _conflict_density(noms_temp)
        max_density_seen = max(max_density_seen, current_density)

        target = collision_target if collision_target is not None else min_conflict_density
        if current_density >= target:
            break
    else:
        if collision_target is not None:
            raise ValueError(
                f"No se pudo alcanzar collision_target={collision_target} "
                f"para n={n}, T={T} despues de 20 intentos. "
                f"Maxima densidad alcanzada: {max_density_seen:.4f}"
            )

    stock_m3  = np.where(
        p_j == 8,
        np.exp(rng.normal(13.0, 0.3, size=n)),
        np.exp(rng.normal(12.5, 0.4, size=n)),
    )
    stock_m3  = np.clip(stock_m3, 150_000, 500_000)
    inflow_m3 = rng.uniform(10_000, 30_000, size=n)
    w_j       = stock_m3 / inflow_m3

    return pd.DataFrame({
        "vessel_id":          [f"V{i+1:02d}" for i in range(n)],
        "r_j":                r_j.astype(int),
        "d_j":                d_j.astype(int),
        "p_j":                p_j.astype(int),
        "w_j":                w_j,
        "stock_acumulado_m3": stock_m3,
        "daily_inflow_m3":    inflow_m3,
    })


print(f"Configuracion:")
print(f"  SCALE_GRID  = {SCALE_GRID}")
print(f"  SEEDS       = {SEEDS}")
print(f"  TIME_LIMIT  = {TIME_LIMIT}s")
print(f"  MIP_GAP     = {MIP_GAP}")
print(f"  RHO_TARGET  = {RHO_TARGET}")
print(f"  Total runs  = {len(SCALE_GRID) * len(SEEDS)}")
print(f"  Output      = {OUTPUT_PATH}")


# CELDA 2: _solve_gurobi_scale(N, seed) -> dict

def _solve_gurobi_scale(N: int, seed: int) -> dict:
    """
    Genera una instancia Scale_N, la resuelve con Gurobi y devuelve un dict de métricas.
    Autocontenido: no depende de estado global (apto para ProcessPoolExecutor).
    """
    import os
    import math
    import time
    import datetime
    import numpy as np
    import pandas as pd

    # Re-cargar .env en el proceso hijo (ProcessPoolExecutor no hereda globals)
    try:
        from pathlib import Path as _Path
        _dotenv = _Path(__file__).resolve().parent.parent.parent / ".env"
        if _dotenv.exists():
            for _line in _dotenv.read_text().splitlines():
                _line = _line.strip()
                if _line and not _line.startswith("#") and "=" in _line:
                    _k, _, _v = _line.partition("=")
                    os.environ.setdefault(_k.strip(), _v.strip())
    except Exception:
        pass

    _RHO_TARGET = 0.80
    _MIX_VLCC   = 0.25
    _R_J_DIST   = "clustered"
    _MIN_CD     = 0.60
    _TIME_LIMIT = 300
    _MIP_GAP    = 1e-4

    # ── helpers locales (proceso hijo no hereda funciones del módulo principal) ──

    def _t_from_rho_local(n: int, vlcc_pct: float, rho_target: float) -> int:
        n_vlcc  = round(n * vlcc_pct)
        n_std   = n - n_vlcc
        total_p = n_vlcc * 8 + n_std * 4
        return math.ceil(total_p / rho_target)

    def _conflict_density_local(nominations_df: pd.DataFrame) -> float:
        r = nominations_df["r_j"].values
        d = nominations_df["d_j"].values
        nn = len(r)
        if nn < 2:
            return 0.0
        total     = nn * (nn - 1) // 2
        conflicts = sum(
            1 for i in range(nn) for j in range(i + 1, nn)
            if r[i] < d[j] and r[j] < d[i]
        )
        return round(conflicts / total, 4)

    def _n_vars_approx_local(n: int, T: int, vlcc_pct: float = 0.25) -> int:
        n_vlcc = round(n * vlcc_pct)
        n_std  = n - n_vlcc
        return (n_std * max(0, T - 4 + 1) + n_vlcc * max(0, T - 8 + 1)) * 2

    def _make_nominations_local(n: int, T: int, vlcc_pct: float,
                                r_j_distribution: str, seed: int,
                                min_conflict_density: float = 0.6) -> pd.DataFrame:
        max_density_seen = 0.0
        for attempt in range(20):
            rng    = np.random.default_rng(seed + attempt)
            n_vlcc = round(n * vlcc_pct)
            n_std  = n - n_vlcc
            p_j    = np.array([8] * n_vlcc + [4] * n_std, dtype=int)
            rng.shuffle(p_j)

            r_max = max(1, T - max(p_j) - 2)

            if r_j_distribution == "clustered":
                n_cluster    = max(1, round(n * 0.7))
                n_spread     = n - n_cluster
                cluster_time = rng.integers(0, max(1, r_max - T // 3))
                spread_times = rng.integers(0, max(1, r_max + 1), size=max(1, n_spread // 2))
                r_cluster    = np.full(n_cluster, cluster_time)
                r_spread     = rng.choice(spread_times, size=n_spread) if n_spread > 0 else np.array([])
                r_j          = np.concatenate([r_cluster, r_spread])
            elif r_j_distribution == "uniform":
                n_batches   = max(1, n // 3)
                batch_times = rng.integers(0, max(1, r_max + 1), size=n_batches)
                r_j         = rng.choice(batch_times, size=n)
            elif r_j_distribution == "random":
                n_batches   = max(1, n // 3)
                batch_times = rng.integers(0, max(1, r_max // 2), size=n_batches)
                r_j         = rng.choice(batch_times, size=n)
            elif r_j_distribution == "bimodal":
                half  = n // 2
                peak1 = max(0, T // 6)
                peak2 = max(0, T // 2)
                r_j   = np.concatenate([np.full(half, peak1), np.full(n - half, peak2)])
            else:
                raise ValueError(f"r_j_distribution desconocida: {r_j_distribution!r}")

            r_j = np.sort(r_j.astype(int))

            t_free    = 0
            fits_in_T = True
            for arr, proc in zip(r_j, p_j):
                start = max(t_free, arr)
                if start + proc > T:
                    fits_in_T = False
                    break
                t_free = start + proc
            if not fits_in_T:
                continue

            slack_vals = rng.integers(1, 4, size=n)
            d_j = np.minimum(r_j + p_j + slack_vals, T)
            d_j = np.maximum(d_j, r_j + p_j)

            noms_temp       = pd.DataFrame({"r_j": r_j, "d_j": d_j, "p_j": p_j})
            current_density = _conflict_density_local(noms_temp)
            max_density_seen = max(max_density_seen, current_density)

            if current_density >= min_conflict_density:
                break

        stock_m3  = np.where(
            p_j == 8,
            np.exp(rng.normal(13.0, 0.3, size=n)),
            np.exp(rng.normal(12.5, 0.4, size=n)),
        )
        stock_m3  = np.clip(stock_m3, 150_000, 500_000)
        inflow_m3 = rng.uniform(10_000, 30_000, size=n)
        w_j       = stock_m3 / inflow_m3

        return pd.DataFrame({
            "vessel_id":          [f"V{i+1:02d}" for i in range(n)],
            "r_j":                r_j.astype(int),
            "d_j":                d_j.astype(int),
            "p_j":                p_j.astype(int),
            "w_j":                w_j,
            "stock_acumulado_m3": stock_m3,
            "daily_inflow_m3":    inflow_m3,
        })

    # ── Generar instancia ──────────────────────────────────────────────────────
    T    = _t_from_rho_local(N, _MIX_VLCC, _RHO_TARGET)
    noms = _make_nominations_local(N, T, _MIX_VLCC, _R_J_DIST, seed, _MIN_CD)

    rho_eff          = float(noms["p_j"].sum()) / T
    collision_density = _conflict_density_local(noms)
    n_vars_qubo      = _n_vars_approx_local(N, T, _MIX_VLCC)

    vessels  = list(noms["vessel_id"])
    r        = {row["vessel_id"]: int(row["r_j"])  for _, row in noms.iterrows()}
    d        = {row["vessel_id"]: int(row["d_j"])  for _, row in noms.iterrows()}
    p        = {row["vessel_id"]: int(row["p_j"])  for _, row in noms.iterrows()}
    w        = {row["vessel_id"]: float(row["w_j"]) for _, row in noms.iterrows()}

    T_viable: dict[str, list[int]] = {
        j: [t for t in range(r[j], T - p[j] + 1)]
        for j in vessels
    }

    # ── Resolver con Gurobi ────────────────────────────────────────────────────
    try:
        import gurobipy as gp
        from gurobipy import GRB

        env = gp.Env(empty=True)
        env.setParam("OutputFlag", 0)
        _wls_id  = os.environ.get("GRB_WLSACCESSID")
        _wls_sec = os.environ.get("GRB_WLSSECRET")
        _wls_lic = os.environ.get("GRB_LICENSEID")
        if _wls_id:
            env.setParam("WLSAccessID", _wls_id)
        if _wls_sec:
            env.setParam("WLSSecret", _wls_sec)
        if _wls_lic:
            env.setParam("LicenseID", int(_wls_lic))
        env.start()

        model = gp.Model(env=env)
        model.Params.TimeLimit = _TIME_LIMIT
        model.Params.MIPGap    = _MIP_GAP
        model.Params.Seed      = seed
        # Threads: todos los disponibles (no se fija)

        x = {
            (j, t): model.addVar(vtype=GRB.BINARY, name=f"x_{j}_{t}")
            for j in vessels
            for t in T_viable[j]
        }
        tard = {j: model.addVar(lb=0.0, name=f"tard_{j}") for j in vessels}
        model.update()

        # (1) Asignación única
        for j in vessels:
            if T_viable[j]:
                model.addConstr(
                    gp.quicksum(x[j, t] for t in T_viable[j]) == 1,
                    name=f"assign_{j}",
                )

        # (2) No solapamiento en pipeline compartido
        for t in range(T):
            occupants = [
                x[j, t2]
                for j in vessels
                for t2 in T_viable[j]
                if t2 <= t < t2 + p[j]
            ]
            if len(occupants) > 1:
                model.addConstr(gp.quicksum(occupants) <= 1, name=f"pipe_{t}")

        # (3) Tardanza linealizada
        for j in vessels:
            if T_viable[j]:
                model.addConstr(
                    tard[j] >= gp.quicksum(
                        (t2 + p[j] - d[j]) * x[j, t2] for t2 in T_viable[j]
                    ),
                    name=f"tard_{j}",
                )

        model.setObjective(
            gp.quicksum(w[j] * tard[j] for j in vessels),
            GRB.MINIMIZE,
        )

        t0 = time.perf_counter()
        model.optimize()
        wall_time = time.perf_counter() - t0

        status_map = {
            GRB.OPTIMAL:    "OPTIMAL",
            GRB.TIME_LIMIT: "TIME_LIMIT",
            GRB.INFEASIBLE: "INFEASIBLE",
            GRB.MEM_LIMIT:  "OOM",
        }
        gurobi_status    = status_map.get(model.Status, f"UNKNOWN_{model.Status}")
        best_obj         = model.ObjVal    if model.SolCount > 0 else float("nan")
        best_bound       = model.ObjBound  if model.SolCount > 0 else float("nan")
        mip_gap_pct      = model.MIPGap * 100.0 if model.SolCount > 0 else float("nan")
        n_nodes_explored = int(model.NodeCount)
        optimality_proven = (
            gurobi_status == "OPTIMAL"
            or (model.SolCount > 0 and model.MIPGap == 0.0)
        )

        model.dispose()
        env.dispose()

    except Exception as exc:
        gurobi_status     = "ERROR"
        best_obj          = float("nan")
        best_bound        = float("nan")
        mip_gap_pct       = float("nan")
        n_nodes_explored  = 0
        optimality_proven = False
        wall_time         = 0.0

    return {
        "exp_id":             "exp07",
        "instance_label":     f"Scale_{N}",
        "N":                  N,
        "T":                  T,
        "rho_target":         _RHO_TARGET,
        "rho_effective":      round(rho_eff, 4),
        "mix_vlcc_pct":       _MIX_VLCC,
        "r_j_distribution":   _R_J_DIST,
        "collision_density":  collision_density,
        "seed":               seed,
        "n_vars_qubo":        n_vars_qubo,
        "gurobi_status":      gurobi_status,
        "best_obj":           best_obj,
        "best_bound":         best_bound,
        "mip_gap_pct":        mip_gap_pct,
        "wall_time_s":        round(wall_time, 3),
        "n_nodes_explored":   n_nodes_explored,
        "optimality_proven":  optimality_proven,
        "run_timestamp":      datetime.datetime.now().isoformat(),
    }


print("_solve_gurobi_scale definida.")


# CELDA 3: RUN — paralelo con ProcessPoolExecutor, append-safe

import concurrent.futures
from concurrent.futures import as_completed
from openpyxl import load_workbook, Workbook

def _load_existing_runs(filepath: Path, sheet: str) -> pd.DataFrame:
    try:
        return pd.read_excel(filepath, sheet_name=sheet)
    except (FileNotFoundError, ValueError):
        return pd.DataFrame()


def _append_rows(filepath: Path, sheet: str, rows: list[dict]) -> None:
    if not rows:
        return
    df_new = pd.DataFrame(rows)
    if not filepath.exists():
        df_new.to_excel(filepath, sheet_name=sheet, index=False)
        return
    wb = load_workbook(filepath)
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
        ws.append(list(df_new.columns))
    else:
        ws = wb[sheet]
    for _, row in df_new.iterrows():
        ws.append(list(row))
    wb.save(filepath)


# Cargar runs ya existentes y filtrar tareas pendientes
existing = _load_existing_runs(OUTPUT_PATH, SHEET)
if not existing.empty:
    done_set = set(zip(existing["N"].astype(int), existing["seed"].astype(int)))
else:
    done_set = set()

tasks_all  = [(N, seed) for N in SCALE_GRID for seed in SEEDS]
tasks_todo = [(N, seed) for N, seed in tasks_all if (N, seed) not in done_set]

print(f"Total tasks: {len(tasks_all)}  |  Ya completadas: {len(done_set)}  |  Pendientes: {len(tasks_todo)}")

if tasks_todo:
    n_workers = min(len(tasks_todo), os.cpu_count() or 1)
    print(f"Lanzando {len(tasks_todo)} runs con {n_workers} workers ...\n")

    with concurrent.futures.ProcessPoolExecutor(max_workers=n_workers) as executor:
        futures = {
            executor.submit(_solve_gurobi_scale, N, seed): (N, seed)
            for N, seed in tasks_todo
        }

        completed_rows: list[dict] = []
        for future in as_completed(futures):
            N, seed = futures[future]
            try:
                row = future.result()
            except Exception as exc:
                row = {
                    "exp_id":             "exp07",
                    "instance_label":     f"Scale_{N}",
                    "N":                  N,
                    "T":                  _t_from_rho(N, MIX_VLCC, RHO_TARGET),
                    "rho_target":         RHO_TARGET,
                    "rho_effective":      float("nan"),
                    "mix_vlcc_pct":       MIX_VLCC,
                    "r_j_distribution":   R_J_DIST,
                    "collision_density":  float("nan"),
                    "seed":               seed,
                    "n_vars_qubo":        _n_vars_approx(N, _t_from_rho(N, MIX_VLCC, RHO_TARGET), MIX_VLCC),
                    "gurobi_status":      "ERROR",
                    "best_obj":           float("nan"),
                    "best_bound":         float("nan"),
                    "mip_gap_pct":        float("nan"),
                    "wall_time_s":        0.0,
                    "n_nodes_explored":   0,
                    "optimality_proven":  False,
                    "run_timestamp":      datetime.datetime.now().isoformat(),
                }

            # Guardar inmediatamente (append-safe)
            _append_rows(OUTPUT_PATH, SHEET, [row])
            completed_rows.append(row)

            done_count = sum(1 for f in futures if f.done())
            gap_str    = f"{row['mip_gap_pct']:.2f}%" if not (
                isinstance(row["mip_gap_pct"], float) and math.isnan(row["mip_gap_pct"])
            ) else "nan%"
            opt_icon   = "✓" if row["optimality_proven"] else "✗"
            print(
                f"  [{done_count:2d}/{len(tasks_todo)}]  "
                f"Scale_{N:<3d}  seed={seed}  "
                f"gap={gap_str:<8}  "
                f"t={row['wall_time_s']:.2f}s  {opt_icon}"
            )

    print(f"\nRuns guardados: {len(completed_rows)}")

# Recargar todos los datos para los análisis siguientes
df_all = _load_existing_runs(OUTPUT_PATH, SHEET)
print(f"Total filas en Excel: {len(df_all)}")

# ── Resumen por N ──────────────────────────────────────────────────────────────
print("\nResumen por N:")
print(f"  {'N':>5}  {'t_mean':>8}  {'t_std':>7}  {'gap_mean':>9}  {'opt':>5}  {'nodes_mean':>12}")
print("  " + "-" * 60)
for N in SCALE_GRID:
    sub = df_all[df_all["N"] == N]
    if sub.empty:
        continue
    t_mean      = sub["wall_time_s"].mean()
    t_std       = sub["wall_time_s"].std(ddof=0)
    gap_vals    = sub["mip_gap_pct"].dropna()
    mean_gap    = gap_vals.mean() if len(gap_vals) > 0 else float("nan")
    n_opt       = int(sub["optimality_proven"].sum())
    mean_nodes  = sub["n_nodes_explored"].mean()
    print(
        f"  Scale_{N:<3d}  "
        f"t={t_mean:.2f}±{t_std:.2f}s  "
        f"gap={mean_gap:.3f}%  "
        f"optimal={n_opt}/{len(sub)}  "
        f"nodes={mean_nodes:.0f}"
    )


# CELDA 4: DETECCIÓN AUTOMÁTICA DE LÍMITES N*, N**, N***

df_all = _load_existing_runs(OUTPUT_PATH, SHEET)

n_star    = None   # primer N con mean(wall_time_s) > 10s
n_dstar   = None   # primer N donde algún seed no certifica óptimo
n_ddstar  = None   # primer N donde ≥3/5 seeds no certifican óptimo

for N in SCALE_GRID:
    sub = df_all[df_all["N"] == N]
    if sub.empty:
        continue

    mean_t  = sub["wall_time_s"].mean()
    n_opt   = int(sub["optimality_proven"].sum())
    n_total = len(sub)
    n_fail  = n_total - n_opt

    if n_star is None and mean_t > 10.0:
        n_star = N

    if n_dstar is None and n_fail >= 1:
        n_dstar = N

    if n_ddstar is None and n_fail >= 3:
        n_ddstar = N

print("\n── Límites de escalabilidad detectados ──────────────────────────────────")
if n_star is not None:
    print(f"  N*   = {n_star:3d}  (primer N con wall_time medio > 10s)")
else:
    print("  N*   = no alcanzado en rango N≤200")

if n_dstar is not None:
    print(f"  N**  = {n_dstar:3d}  (primer N donde algún seed no certifica óptimo)")
else:
    print("  N**  = no alcanzado en rango N≤200")

if n_ddstar is not None:
    print(f"  N*** = {n_ddstar:3d}  (primer N donde ≥3/5 seeds no certifican óptimo)")
else:
    print("  N*** = no alcanzado en rango N≤200")
print()


# CELDA 5: PLOTS

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

df_all = _load_existing_runs(OUTPUT_PATH, SHEET)

# Pre-computar agregados por N
agg: dict[int, dict] = {}
for N in SCALE_GRID:
    sub = df_all[df_all["N"] == N]
    if sub.empty:
        continue
    n_opt   = int(sub["optimality_proven"].sum())
    n_total = len(sub)
    gap_vals = sub["mip_gap_pct"].dropna()
    n_vars   = sub["n_vars_qubo"].iloc[0]
    agg[N] = {
        "t_mean":       sub["wall_time_s"].mean(),
        "t_std":        sub["wall_time_s"].std(ddof=0),
        "gap_mean":     gap_vals.mean() if len(gap_vals) > 0 else float("nan"),
        "gap_std":      gap_vals.std(ddof=0) if len(gap_vals) > 1 else 0.0,
        "n_opt":        n_opt,
        "n_total":      n_total,
        "nodes_mean":   sub["n_nodes_explored"].mean(),
        "nodes_std":    sub["n_nodes_explored"].std(ddof=0),
        "n_vars":       n_vars,
    }


def _color_for_N(N: int, data: dict) -> str:
    n_opt   = data["n_opt"]
    n_total = data["n_total"]
    if n_opt == n_total:
        return "#2ca02c"   # verde — todos certifican óptimo
    if n_opt == 0:
        return "#d62728"   # rojo — ninguno certifica
    return "#ff7f0e"       # naranja — parcial


Ns_agg  = [N for N in SCALE_GRID if N in agg]
colors  = [_color_for_N(N, agg[N]) for N in Ns_agg]

# ── Plot 1: wall_time scaling ─────────────────────────────────────────────────
fig, ax1 = plt.subplots(figsize=(11, 6))

ax2 = ax1.twinx()
ax2.plot(
    Ns_agg,
    [agg[N]["n_vars"] for N in Ns_agg],
    color="silver", linewidth=1.2, linestyle="--", label="n_vars_qubo",
)
ax2.set_ylabel("n_vars_qubo (aprox.)", color="grey", fontsize=10)
ax2.tick_params(axis="y", labelcolor="grey")

ax1.set_yscale("log")
for N, col in zip(Ns_agg, colors):
    ax1.errorbar(
        N, agg[N]["t_mean"],
        yerr=agg[N]["t_std"],
        fmt="o", color=col, capsize=4, markersize=6,
    )
ax1.plot(Ns_agg, [agg[N]["t_mean"] for N in Ns_agg],
         color="steelblue", linewidth=1.0, alpha=0.5)

# Líneas verticales en N*, N**, N***
for val, label_str, ls in [
    (n_star,   "N*",   "--"),
    (n_dstar,  "N**",  "-."),
    (n_ddstar, "N***", ":"),
]:
    if val is not None:
        ax1.axvline(val, linestyle=ls, color="black", linewidth=1.0, alpha=0.7)
        ax1.text(val + 1, ax1.get_ylim()[0] * 1.5, label_str, fontsize=9)

ax1.set_xlabel("N (número de buques)", fontsize=11)
ax1.set_ylabel("wall_time_s (log)", fontsize=11)
ax1.set_title("Gurobi scalability — offshore BAP time-indexed\nwall time vs N", fontsize=12)

from matplotlib.patches import Patch
legend_elements = [
    Patch(facecolor="#2ca02c", label="todos certifican óptimo"),
    Patch(facecolor="#ff7f0e", label="algún seed no certifica"),
    Patch(facecolor="#d62728", label="ninguno certifica óptimo"),
    plt.Line2D([0], [0], color="silver", linestyle="--", label="n_vars_qubo"),
]
ax1.legend(handles=legend_elements, loc="upper left", fontsize=9)
ax1.grid(True, which="both", linestyle=":", alpha=0.4)

plt.tight_layout()
out1 = RESULTS_DIR / "exp07_walltime_scaling.png"
fig.savefig(out1, dpi=150)
plt.close(fig)
print(f"Guardado: {out1}")

# ── Plot 2: MIP gap scaling ────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(11, 5))

ax.axhline(0.0, color="black", linewidth=0.8, linestyle="--", label="óptimo certificado (0%)")
for N, col in zip(Ns_agg, colors):
    ax.errorbar(
        N, agg[N]["gap_mean"],
        yerr=agg[N]["gap_std"] if not math.isnan(agg[N]["gap_std"]) else 0,
        fmt="o", color=col, capsize=4, markersize=6,
    )
ax.plot(Ns_agg, [agg[N]["gap_mean"] for N in Ns_agg],
        color="steelblue", linewidth=1.0, alpha=0.5)

for val, label_str, ls in [
    (n_star,   "N*",   "--"),
    (n_dstar,  "N**",  "-."),
    (n_ddstar, "N***", ":"),
]:
    if val is not None:
        ax.axvline(val, linestyle=ls, color="black", linewidth=1.0, alpha=0.7)
        ax.text(val + 1, ax.get_ylim()[1] * 0.9 if ax.get_ylim()[1] > 0 else 0.01,
                label_str, fontsize=9)

ax.set_xlabel("N (número de buques)", fontsize=11)
ax.set_ylabel("MIP gap al terminar (%)", fontsize=11)
ax.set_title("Gurobi scalability — MIP gap vs N", fontsize=12)
ax.legend(handles=legend_elements[:3], loc="upper left", fontsize=9)
ax.grid(True, linestyle=":", alpha=0.4)

plt.tight_layout()
out2 = RESULTS_DIR / "exp07_mipgap_scaling.png"
fig.savefig(out2, dpi=150)
plt.close(fig)
print(f"Guardado: {out2}")

# ── Plot 3: nodos explorados ────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(11, 5))

nodes_means = [agg[N]["nodes_mean"] for N in Ns_agg]
nodes_stds  = [agg[N]["nodes_std"]  for N in Ns_agg]

# Usar escala log solo si hay valores > 0
positive_nodes = [v for v in nodes_means if v > 0]
if positive_nodes:
    ax.set_yscale("log")

for N, col, nm, ns in zip(Ns_agg, colors, nodes_means, nodes_stds):
    ax.errorbar(N, max(nm, 1), yerr=ns if ns > 0 else None,
                fmt="o", color=col, capsize=4, markersize=6)
ax.plot(Ns_agg, [max(v, 1) for v in nodes_means],
        color="steelblue", linewidth=1.0, alpha=0.5)

for val, label_str, ls in [
    (n_star,   "N*",   "--"),
    (n_dstar,  "N**",  "-."),
    (n_ddstar, "N***", ":"),
]:
    if val is not None:
        ax.axvline(val, linestyle=ls, color="black", linewidth=1.0, alpha=0.7)

ax.set_xlabel("N (número de buques)", fontsize=11)
ax.set_ylabel("nodos B&B explorados (log)", fontsize=11)
ax.set_title("Gurobi scalability — nodos B&B vs N\n(presolve trivial → B&B = 0)", fontsize=12)
ax.legend(handles=legend_elements[:3], loc="upper left", fontsize=9)
ax.grid(True, which="both", linestyle=":", alpha=0.4)

plt.tight_layout()
out3 = RESULTS_DIR / "exp07_nodes_scaling.png"
fig.savefig(out3, dpi=150)
plt.close(fig)
print(f"Guardado: {out3}")


# CELDA 6: RESUMEN TEXTUAL + exp07_summary.csv

df_all = _load_existing_runs(OUTPUT_PATH, SHEET)

summary_rows = []
for N in SCALE_GRID:
    sub = df_all[df_all["N"] == N]
    if sub.empty:
        continue
    n_opt    = int(sub["optimality_proven"].sum())
    n_total  = len(sub)
    gap_vals = sub["mip_gap_pct"].dropna()
    T_val    = int(sub["T"].iloc[0])
    n_vars   = int(sub["n_vars_qubo"].iloc[0])
    summary_rows.append({
        "N":               N,
        "T":               T_val,
        "n_vars_qubo":     n_vars,
        "rho_target":      RHO_TARGET,
        "rho_eff_mean":    round(sub["rho_effective"].mean(), 4),
        "cd_mean":         round(sub["collision_density"].mean(), 4),
        "t_mean_s":        round(sub["wall_time_s"].mean(), 3),
        "t_std_s":         round(sub["wall_time_s"].std(ddof=0), 3),
        "gap_mean_pct":    round(gap_vals.mean(), 4) if len(gap_vals) > 0 else float("nan"),
        "gap_std_pct":     round(gap_vals.std(ddof=0), 4) if len(gap_vals) > 1 else float("nan"),
        "n_optimal":       n_opt,
        "n_seeds":         n_total,
        "nodes_mean":      round(sub["n_nodes_explored"].mean(), 1),
        "nodes_std":       round(sub["n_nodes_explored"].std(ddof=0), 1),
        "is_n_star":       N == n_star,
        "is_n_dstar":      N == n_dstar,
        "is_n_ddstar":     N == n_ddstar,
    })

df_summary = pd.DataFrame(summary_rows)
csv_path   = RESULTS_DIR / "exp07_summary.csv"
df_summary.to_csv(csv_path, index=False)
print(f"Guardado: {csv_path}\n")

# Tabla formateada en consola
header = (
    f"  {'N':>5}  {'T':>4}  {'n_vars':>8}  "
    f"{'t_mean':>8}  {'t_std':>7}  "
    f"{'gap_mean':>9}  {'opt':>6}  {'nodes':>10}"
)
print("── Resumen Exp 07 ─────────────────────────────────────────────────────────────")
print(header)
print("  " + "─" * (len(header) - 2))
for row in summary_rows:
    gap_str   = f"{row['gap_mean_pct']:.4f}%" if not math.isnan(row["gap_mean_pct"]) else "   nan%"
    flags     = ""
    if row["is_n_star"]:   flags += " ←N*"
    if row["is_n_dstar"]:  flags += " ←N**"
    if row["is_n_ddstar"]: flags += " ←N***"
    print(
        f"  Scale_{row['N']:<3d}  "
        f"T={row['T']:<4d}  "
        f"n_vars={row['n_vars_qubo']:<8d}  "
        f"t={row['t_mean_s']:.2f}±{row['t_std_s']:.2f}s  "
        f"gap={gap_str:<10}  "
        f"opt={row['n_optimal']}/{row['n_seeds']}  "
        f"nodes={row['nodes_mean']:.0f}"
        f"{flags}"
    )

print()
print("── Límites ────────────────────────────────────────────────────────────────────")
print(f"  N*   = {n_star   if n_star   is not None else 'no alcanzado (N≤200)'}")
print(f"  N**  = {n_dstar  if n_dstar  is not None else 'no alcanzado (N≤200)'}")
print(f"  N*** = {n_ddstar if n_ddstar is not None else 'no alcanzado (N≤200)'}")
print()
print(f"Output principal : {OUTPUT_PATH}")
print(f"Resumen CSV      : {csv_path}")
print(f"Plots            : {RESULTS_DIR}/exp07_*.png")
