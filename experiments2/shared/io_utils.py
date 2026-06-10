"""
io_utils.py — Excel I/O utilities for instances and results.

Handles:
  - Saving generated instances to data/instances.xlsx (hojas: size, dens, slack)
  - Loading instances from Excel
  - Append-safe result writing (load_existing_runs, append_rows)
  - Metadata via hoja 'metadata' (save_metadata, load_metadata)
"""

import io
import subprocess
import logging
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# Paths
EXPERIMENTS2_DIR = Path(__file__).parent.parent
DATA_DIR         = EXPERIMENTS2_DIR / "data"
RESULTS_DIR      = EXPERIMENTS2_DIR / "results"
INSTANCES_EXCEL  = DATA_DIR / "instances.xlsx"


def ensure_directories() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    logger.info("Directories ensured.")


def get_commit_hash() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "HEAD"],
            cwd=str(EXPERIMENTS2_DIR.parent),
            stderr=subprocess.DEVNULL,
        ).decode().strip()
    except Exception:
        return "unknown"


# ---------------------------------------------------------------------------
# Instances Excel (una vez, generadas en setup.py)
# ---------------------------------------------------------------------------

def save_instances_to_excel(instances: dict) -> None:
    """
    Guarda instancias en data/instances.xlsx.

    Args:
        instances: {"size": {label: inst_dict}, "dens": {...}, "slack": {...}}
            Cada inst_dict debe tener "nominations" (DataFrame) + metadatos.
    """
    ensure_directories()
    with pd.ExcelWriter(INSTANCES_EXCEL, engine="openpyxl") as writer:
        for axis, axis_dict in instances.items():
            rows = []
            for label, inst in axis_dict.items():
                noms = inst["nominations"]
                row = {
                    "instance_label":  label,
                    "N":               inst["N"],
                    "T":               inst["T"],
                    "rho_target":      inst.get("rho_target", float("nan")),
                    "rho_effective":   inst.get("rho_effective", float("nan")),
                    "mix_vlcc_pct":    inst.get("mix_vlcc_pct", float("nan")),
                    "r_j_distribution": inst.get("r_j_distribution", "uniform"),
                    "collision_target":  inst.get("collision_target", float("nan")),
                    "collision_density": inst.get("collision_density", float("nan")),
                    "nominations_json": noms.to_json(orient="records"),
                }
                rows.append(row)
            pd.DataFrame(rows).to_excel(writer, sheet_name=axis, index=False)
    logger.info("Instances saved: %s", INSTANCES_EXCEL)


def load_instances_from_excel(axis: str) -> dict:
    """
    Carga instancias de data/instances.xlsx para el eje indicado.

    Returns:
        {instance_label: inst_dict}  — inst_dict incluye "nominations" como DataFrame.
    """
    if not INSTANCES_EXCEL.exists():
        raise FileNotFoundError(f"Instances Excel not found: {INSTANCES_EXCEL}")

    df = pd.read_excel(INSTANCES_EXCEL, sheet_name=axis)
    result = {}
    for _, row in df.iterrows():
        label = row["instance_label"]
        noms  = pd.read_json(io.StringIO(row["nominations_json"]), orient="records")
        result[label] = {
            "instance_label":  label,
            "N":               int(row["N"]),
            "T":               int(row["T"]),
            "rho_target":      float(row["rho_target"]),
            "rho_effective":   float(row["rho_effective"]),
            "mix_vlcc_pct":    float(row.get("mix_vlcc_pct", float("nan"))),
            "r_j_distribution": str(row.get("r_j_distribution", "uniform")),
            "nominations":     noms,
        }
    return result


# ---------------------------------------------------------------------------
# Append-safe result writing
# ---------------------------------------------------------------------------

def load_existing_runs(filepath: str | Path, sheet: str) -> pd.DataFrame:
    """Carga runs existentes. Devuelve DataFrame vacío si no existe el archivo/hoja."""
    try:
        return pd.read_excel(filepath, sheet_name=sheet)
    except (FileNotFoundError, ValueError):
        return pd.DataFrame()


def append_rows(filepath: str | Path, sheet: str, rows: list[dict]) -> None:
    """Appenda filas al Excel. Crea el archivo/hoja si no existe."""
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    df_new = pd.DataFrame(rows)
    if not filepath.exists():
        df_new.to_excel(filepath, sheet_name=sheet, index=False)
        logger.info("Created %s (sheet=%s, %d rows)", filepath.name, sheet, len(df_new))
        return

    wb = openpyxl.load_workbook(filepath)
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
        ws.append(list(df_new.columns))
    else:
        ws = wb[sheet]

    for _, row in df_new.iterrows():
        ws.append([_safe_val(v) for v in row])
    wb.save(filepath)
    logger.info("Appended %d rows → %s sheet=%s", len(df_new), filepath.name, sheet)


def _safe_val(v):
    """Convierte numpy scalars y NaN a tipos nativos para openpyxl."""
    if isinstance(v, float) and np.isnan(v):
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    if isinstance(v, (np.bool_,)):
        return bool(v)
    return v


# ---------------------------------------------------------------------------
# Metadata — hoja 'metadata' en cada Excel (no custom doc props)
# ---------------------------------------------------------------------------

def save_metadata(filepath: str | Path, params: dict) -> None:
    """Escribe/actualiza la hoja 'metadata' en el Excel."""
    filepath = Path(filepath)
    if not filepath.exists():
        logger.warning("save_metadata: file does not exist yet: %s", filepath)
        return

    wb = openpyxl.load_workbook(filepath)
    if "metadata" not in wb.sheetnames:
        ws = wb.create_sheet("metadata")
        ws.append(["key", "value", "updated_at"])
    else:
        ws = wb["metadata"]

    now      = datetime.now().isoformat()
    existing = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
    for key, value in params.items():
        if key in existing:
            ws.cell(existing[key], 2).value = str(value)
            ws.cell(existing[key], 3).value = now
        else:
            ws.append([key, str(value), now])
    wb.save(filepath)
    logger.info("Metadata saved to %s: %s", filepath.name, list(params.keys()))


def load_metadata(filepath: str | Path) -> dict:
    """Lee la hoja 'metadata' y devuelve un dict key→value."""
    filepath = Path(filepath)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        if "metadata" not in wb.sheetnames:
            return {}
        ws = wb["metadata"]
        return {
            ws.cell(r, 1).value: ws.cell(r, 2).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value
        }
    except FileNotFoundError:
        return {}


# ---------------------------------------------------------------------------
# D-Wave / solver timing extraction
# ---------------------------------------------------------------------------

def extract_solver_timing(sampleset) -> dict:
    """
    Extrae todos los campos de timing disponibles en un SampleSet de D-Wave.

    Para QPU directo (EmbeddingComposite + DWaveSampler):
      - qpu_sampling_time_us     : tiempo puro de QPU (annealing + readout), en µs
      - qpu_anneal_time_per_sample_us : tiempo de annealing por muestra, en µs
      - qpu_readout_time_per_sample_us: tiempo de readout por muestra, en µs
      - qpu_access_time_us       : acceso total al QPU sin latencia de red, en µs
      - qpu_access_overhead_time_us : overhead de embedding/programación, en µs
      - total_post_processing_time_us: post-procesamiento en servidor D-Wave, en µs

    Para LeapHybridSampler:
      - lh_run_time_us           : cómputo real del solver híbrido sin latencia, en µs
      - lh_run_time_s            : idem en segundos (para comparar con wall_time_s)
      - n_solver_calls           : iteraciones internas del solver híbrido

    Para SimulatedAnnealingSampler:
      - sa_timing_us             : tiempo de cómputo interno si está disponible

    Todos los campos no disponibles se rellenan con float("nan").
    El wall_time_s debe medirse siempre con perf_counter() y reportarse por separado
    como referencia del tiempo total incluyendo latencia de red.
    """
    info = getattr(sampleset, "info", {}) or {}
    timing = info.get("timing", {}) or {}

    def _us(key: str) -> float:
        v = timing.get(key)
        return float(v) if v is not None else float("nan")

    def _info(key: str) -> float:
        v = info.get(key)
        return float(v) if v is not None else float("nan")

    # QPU timing (DWaveSampler vía EmbeddingComposite)
    qpu_sampling        = _us("qpu_sampling_time")
    qpu_anneal_per      = _us("qpu_anneal_time_per_sample")
    qpu_readout_per     = _us("qpu_readout_time_per_sample")
    qpu_access          = _us("qpu_access_time")
    qpu_access_overhead = _us("qpu_access_overhead_time")
    qpu_post_proc       = _us("total_post_processing_time")

    # LeapHybrid timing — run_time está en µs en sampleset.info (no en timing)
    lh_run_time_us = _info("run_time")
    lh_run_time_s  = lh_run_time_us / 1_000_000.0 if not (lh_run_time_us != lh_run_time_us) else float("nan")
    n_solver_calls = _info("n_solver_calls")

    # SA timing — disponible si el sampler lo reporta
    sa_timing_us = _us("sampling_time") if not timing else float("nan")

    return {
        "qpu_sampling_time_us":          qpu_sampling,
        "qpu_anneal_time_per_sample_us": qpu_anneal_per,
        "qpu_readout_time_per_sample_us": qpu_readout_per,
        "qpu_access_time_us":            qpu_access,
        "qpu_access_overhead_time_us":   qpu_access_overhead,
        "total_post_processing_time_us": qpu_post_proc,
        "lh_run_time_us":                lh_run_time_us,
        "lh_run_time_s":                 round(lh_run_time_s, 3) if lh_run_time_s == lh_run_time_s else float("nan"),
        "n_solver_calls":                n_solver_calls,
        "sa_timing_us":                  sa_timing_us,
    }


# ---------------------------------------------------------------------------
# Compat: funciones usadas por scripts más viejos (mantener firma)
# ---------------------------------------------------------------------------

def ensure_excel_exists(filepath: Path, sheets_dict: dict) -> None:
    filepath.parent.mkdir(parents=True, exist_ok=True)
    if filepath.exists():
        return
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, columns in sheets_dict.items():
            pd.DataFrame(columns=columns).to_excel(writer, sheet_name=sheet_name, index=False)
    logger.info("Created Excel: %s", filepath)


def append_to_excel(filepath: Path, sheet_name: str, df: pd.DataFrame) -> None:
    append_rows(filepath, sheet_name, df.to_dict(orient="records"))


def add_metadata_to_excel(filepath: Path, **metadata) -> None:
    save_metadata(filepath, metadata)


def get_solver_dir(solver: str) -> Path:
    return RESULTS_DIR
